from django.db.models import Avg, Count, Q, Sum, Max
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter, SearchFilter
from .models import Run, Position, Subscription, Challenge, AthleteInfo, CollectibleItem
from .serializers import RunSerializer, PositionSerializer, UserSerializer, AthleteSerializer, CoachSerializer, \
    ChallengeSerializer, CollectibleItemSerializer
from geopy.distance import geodesic
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth.models import User
from django.db import connection
import openpyxl as op


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'id']
    ordering_fields = ['created_at']


class StatusStartView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Ты чо! Этот забег стартовать нельзя, он уже стартовал'},
                            status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def status_stop_view(request, run_id):
    run = get_object_or_404(Run,id=run_id)
    if run.status == 'in_progress':
        run.status = 'finished'
        #Добавлю условие чтобы не вылетало исключение когда делаю задачу которая идет ДО создания позиций
        #А МОЖНО ПРОСТО ВРЕМЕННО ЗАКОММЕНТИРОВАТЬ ВСЕ ДО run.save()
        if Position.objects.filter(run=run_id).exists():
            positions_qs = Position.objects.filter(run=run_id)
            positions_quantity = len(positions_qs)
            distance = 0
            for i in range(positions_quantity-1):
                distance += geodesic((positions_qs[i].latitude,positions_qs[i].longitude), (positions_qs[i+1].latitude,positions_qs[i+1].longitude)).kilometers
            run.distance = distance
            # -----------------------------------------
            positions_qs_sorted_by_date = positions_qs.order_by('date_time')
            run_time = positions_qs_sorted_by_date[positions_quantity-1].date_time-positions_qs_sorted_by_date[0].date_time
            run.run_time_seconds = run_time.total_seconds()
            #-------------------------------------------
            average_speed = positions_qs.aggregate(Avg('speed'))
            run.speed = round(average_speed['speed__avg'], 2)
        run.save()
        #-------------------------------------------
        if Run.objects.filter(status='finished', athlete=run.athlete).count() >= 10:
            challenge, created = Challenge.objects.get_or_create(full_name = 'Сделай 10 Забегов!', athlete=run.athlete)
        #-------------------------------------------
        distance_sum = Run.objects.filter(status='finished', athlete=run.athlete).aggregate(Sum('distance'))

        if distance_sum['distance__sum'] >= 50:
            challenge, created = Challenge.objects.get_or_create(full_name = 'Пробеги 50 километров!', athlete=run.athlete)
        #-------------------------------------------
        if run.distance >= 2 and run.run_time_seconds <= 600:
            challenge, created = Challenge.objects.get_or_create(full_name='Пробеги 2 километра меньше чем за 10 минут!',
                                                                 athlete=run.athlete)
        return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)
    else:
        return Response({'message': 'Ты чо! Этот забег финишировать нельзя, он еще не стартовал или уже завершен'},
                        status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def company_details(request):
    return Response({'company_name': 'Космическая Картошка',
                     'slogan':'Картофельное Пюре - путь к звездам!',
                     'contacts': 'Город Задунайск, улица Картофельная дом 16'}, status=status.HTTP_200_OK)


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = self.queryset
        run_id = self.request.query_params.get('run')
        if run_id:
            qs = qs.filter(run=run_id)
        return qs

    def perform_create(self, serializer):
        run = serializer.validated_data['run']
        serializer.save() # чтобы включить только что создаваемую позицию в QS
        all_positions = Position.objects.filter(run=run)
        if all_positions.count() > 1:

            ordered_positions = all_positions.order_by('-id')
            last_position = ordered_positions[0]
            previous_position = ordered_positions[1]
            previous_distance = previous_position.distance
            last_distance = geodesic((last_position.latitude, last_position.longitude),
                                 (previous_position.latitude, previous_position.longitude)).meters
            time_delta = last_position.date_time - previous_position.date_time
            speed = last_distance/time_delta.total_seconds()
            last_position.speed = round(speed, 2)
            last_position.distance = round(previous_distance + last_distance/1000, 2)
            last_position.save()
        # if расстояние между этой позицией и каким то из коллектибле итемз < 100 метров, то:
            # ItemAthletRelation.objects.create(athlete=, item=)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserSerializer
    queryset = User.objects.filter(is_superuser=False)
    filter_backends = [SearchFilter]
    search_fields = ['first_name', 'last_name']

    def get_queryset(self):

        qs = self.queryset
        user_type = self.request.query_params.get('type')
        if user_type and user_type=='coach':
            qs = qs.filter(is_staff=True)
        if user_type and user_type=='athlete':
            qs = qs.filter(is_staff=False)
        qs = qs.annotate(runs_finished=Count('run', filter=Q(run__status='finished')))
        qs = qs.annotate(rating=Avg('athletes__rating')) # Вариант №1, поле rating вычисляется и добавляется здесь

        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            user = self.get_object() # крутой метод! почему я его не знал!!
            if user.is_staff:
                return CoachSerializer
            else:
                return AthleteSerializer
        return super().get_serializer_class()


class SubscribeView(APIView):
    def post(self, request, id):
        coach_id = id
        athlete_id = self.request.data['athlete']

        coach = get_object_or_404(User, id=coach_id)
        if not User.objects.filter(id=athlete_id).exists():
            return Response({'message': f'Пользователя c id {athlete_id} не существует'},
                            status=status.HTTP_400_BAD_REQUEST)
        athlete=User.objects.get(id=athlete_id)

        if not coach.is_staff:
            return Response({'message': f'Пользователь c id {coach_id} это не тренер'}, status=status.HTTP_400_BAD_REQUEST)
        if athlete.is_staff:
            return Response({'message': f'Пользователь c id {coach_id} это не бегун'}, status=status.HTTP_400_BAD_REQUEST)
        if Subscription.objects.filter(coach=coach, athlete=athlete).exists():
            return Response({'message': 'Такая подписка уже существует'},
                            status=status.HTTP_400_BAD_REQUEST)
        Subscription.objects.create(coach=coach, athlete=athlete)

        return Response({'message': 'Все ништяк'}, status=status.HTTP_200_OK)


class ChallengeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['athlete']


@api_view(['GET'])
def challenge_summary_view(request):
    # Получим список уникальных названий челленджей
    unique_challenges = Challenge.objects.values('full_name').distinct()
    unique_name_list = [i['full_name'] for i in unique_challenges]

    #Получим список атлетов для каждого челленджа
    final_list = []
    for challenge_name in unique_name_list:
        user_queryset = User.objects.filter(challenge__full_name=challenge_name)
        athletes_list = []
        for athlete in user_queryset:
            athletes_list.append({'id': athlete.id, 'full_name': f'{athlete.first_name} {athlete.last_name}'})

        final_list.append({'name_to_display':challenge_name, 'athletes':athletes_list})

    return Response(final_list)


class CoachRatingView(APIView):
    def post(self, request, coach_id):
        athlete_id = request.data.get('athlete')
        rating = request.data.get('rating')

        user_coach = get_object_or_404(User, id=coach_id)

        if not athlete_id:
            return Response({'message': f'Нет поля athlete_id'}, status=status.HTTP_400_BAD_REQUEST)

        if not rating:
            return Response({'message': f'Нет поля rating'}, status=status.HTTP_400_BAD_REQUEST)

        if isinstance(rating, str) and not rating.isdigit():
            return Response({'message': f'У вас rating не цифра. Ваше значение {rating}'},
                            status=status.HTTP_400_BAD_REQUEST)

        if not 1<= int(rating) <=5:
            return Response({'message': f'rating не в пределах от 1 до 5. Ваше значение {rating}'}, status=status.HTTP_400_BAD_REQUEST)

        user_athlete = get_object_or_404(User, id=athlete_id)

        if Subscription.objects.filter(coach=coach_id, athlete=athlete_id).exists():
            subscription = Subscription.objects.get(coach=coach_id, athlete=athlete_id)
            subscription.rating = rating
            subscription.save()
        else:
            return Response({'message':f'Бегун c id {athlete_id} не подписан на тренера с id {coach_id}'},
                            status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'ОК'})


class AnalyticsCoachView(APIView):
    def get(self, request, coach_id):
        coach_queryset = Subscription.objects.filter(coach=coach_id)

        qs_with_additional_fields = coach_queryset.annotate(
            max_distance=Max('athlete__run__distance'),
            sum_distances=Sum('athlete__run__distance'),
            avg_speed=Avg('athlete__run__speed')
        )

        longest_qs = qs_with_additional_fields.order_by('-max_distance').first()
        longest_run_value = longest_qs.max_distance
        longest_run_user = longest_qs.athlete_id

        max_total_run_qs = qs_with_additional_fields.order_by('-sum_distances').first()
        total_run_value = max_total_run_qs.sum_distances
        total_run_user = max_total_run_qs.athlete_id

        max_avg_speed_qs = qs_with_additional_fields.order_by('-avg_speed').first()
        speed_avg_value = max_avg_speed_qs.avg_speed
        speed_avg_user = max_avg_speed_qs.athlete_id


        return Response({'longest_run_value': longest_run_value,
                         'longest_run_user': longest_run_user,
                         'total_run_value': total_run_value,
                         'total_run_user': total_run_user,
                         'speed_avg_value': speed_avg_value,
                         'speed_avg_user': speed_avg_user
                         })


class AthleteInfoView(APIView):
    def get(self, request, user_id):
        user = User.objects.get(id=user_id)
        athlete_info, created = AthleteInfo.objects.get_or_create(user=user)
        return Response({'level': athlete_info.level,
                         'goals': athlete_info.goals,
                         'user_id': athlete_info.user.id

        })

    def put(self, request, user_id):
        goals = request.data.get('goals')
        level =  request.data.get('level')

        if not str(level).isdigit() or int(level) not in range(1,6):
            return Response({'message': 'level должен быть числом от 1 до 5'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(id=user_id).exists():
            user = User.objects.get(id=user_id)
        else:
            return Response({'message':'Такого User не существует'}, status=status.HTTP_404_NOT_FOUND)

        athlete_info, created = AthleteInfo.objects.update_or_create(
            user = user,
            defaults = {'goals':goals, 'level':level}
        )

        return Response({'message': 'Создано/изменено'}, status=status.HTTP_201_CREATED)

@api_view(['POST'])
def upload_view(request):
    if request.method == 'POST' and request.FILES.get('file'):
    # if True:
        uploaded_xlsx_file = request.FILES['file']
        # wb =  op.load_workbook('upload_example.xlsx')
        wb = op.load_workbook(uploaded_xlsx_file, data_only=True)
        sheet = wb.active
        wrong_rows_list=[]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            name, uid, value, latitude, longitude, picture = row
            data = {
                'name': name,
                'uid': uid,
                'latitude': latitude,
                'longitude': longitude,
                'picture': picture,
                'value': value,
            }
            serializer = CollectibleItemSerializer(data=data)
            if serializer.is_valid():
                CollectibleItem.objects.create(name=name,
                                               uid=uid,
                                               value=value,
                                               latitude=latitude,
                                               longitude=longitude,
                                               picture=picture)
            else:
                wrong_rows_list.append([name,uid,value,latitude,longitude,picture])

        print(wrong_rows_list)
        return Response(wrong_rows_list)
    return Response([])


class CollectibleItemViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer

    def get_queryset(self):
        qs = self.queryset
        for i in qs:
            print(i)
        return qs